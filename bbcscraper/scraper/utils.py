from openpyxl import load_workbook
import os

def add_data_to_excel_file(data):
    file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'BBCNews.xlsx')
    wb = load_workbook(file_path)
    ws = wb.active
    if ws.max_row == 1:
        ws.append(["Link", "Title", "Description"])
    if data['url'] in [cell.value for cell in ws['A']]:
        return
    else:
        ws.append([data['url'], data['title'], data['description']])
    wb.save(file_path)